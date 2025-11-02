import argparse
import os
import pandas as pd
from openpyxl import Workbook
from Tool import StoredResult

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--routeid", type=int)
    parser.add_argument("--direction", type=int)
    parser.add_argument("--test_date", type=str, default="2025-09-23")
    parser.add_argument("--day", type=str)
    parser.add_argument("--mode", type=str)

    opt = parser.parse_args()
    print(opt)

    ROUTEID, DIRECTION, DAY, MODE = opt.routeid, opt.direction, opt.day, opt.mode
    TEST_DATE = opt.test_date.replace("-", "")
    total_1, correct_10_1, correct_30_1, correct_60_1, correct_120_1 = 0, 0, 0, 0, 0
    total_2, correct_10_2, correct_30_2, correct_60_2, correct_120_2 = 0, 0, 0, 0, 0
    total_3, correct_10_3, correct_30_3, correct_60_3, correct_120_3 = 0, 0, 0, 0, 0

    """
    建置對應的存放路經，方便後續檔案存放
    """
    if not os.path.exists(f"inference_result/{ROUTEID}/{DIRECTION}/{MODE}/"):
        os.makedirs(f"inference_result/{ROUTEID}/{DIRECTION}/{MODE}/")

        """
    初始化一個Excel檔案，用來存放結果
    """
    ResultWB = Workbook()

    """
    根據每一站去進行預測
    """
    sheets = pd.ExcelFile(f"training_dataset/{ROUTEID}/{DIRECTION}/dataset_{DAY}.xlsx").sheet_names
    ResultWB = Workbook()

    for sheet in sheets[1:]:
        """
        初始化一個excel檔案，用於進行資料新增
        """
        ResultWS = ResultWB.create_sheet(title=sheet)
        ResultWS.append(["Date", "SechudeleIndex", "h_driveTime", "Ratio", "std", "h_stayTime", "c_stayTime", "GroundTruth", "Predicted"])
        ResultWS.column_dimensions["A"].width = 20
        ResultWS.column_dimensions["B"].width = 20
        ResultWS.column_dimensions["C"].width = 20
        ResultWS.column_dimensions["D"].width = 20
        ResultWS.column_dimensions["E"].width = 20
        ResultWS.column_dimensions["F"].width = 20
        ResultWS.column_dimensions["G"].width = 20
        ResultWS.column_dimensions["H"].width = 20
        ResultWS.column_dimensions["I"].width = 20

        for PredictedType in ["-1", "-2", "-other"]:
            """
            抓取對應參數
            """
            ParameterDF = pd.read_excel(f"./training_result/{ROUTEID}/{DIRECTION}/{MODE}/parameters_{DAY}.xlsx")
            Query = (ParameterDF["stop"].str.contains(sheet)) & (ParameterDF["stop"].str.contains(PredictedType))
            ParameterDF = ParameterDF[Query]
            if len(ParameterDF) == 0:
                continue

            DatasetDF = pd.read_excel(f"training_dataset/{ROUTEID}/{DIRECTION}/dataset_{DAY}.xlsx", sheet_name=sheet)
            Query = (DatasetDF["Date"] == int(TEST_DATE)) & (DatasetDF["SechudeleIndex"].str.contains(PredictedType))
            DatasetDF = DatasetDF[Query]

            """
            根據不同模式進行預測
            """
            for index in range(len(DatasetDF)):
                Data = DatasetDF.iloc[index]
                HDriveTime, HStayTime, CStayTime, Ratio, std, avg, GroundTruth = Data["h_driveTime"], Data["h_stayTime"], Data["c_stayTime"], Data["Ratio"], Data["std"], Data["avg"], Data["GroundTruth"]
                if MODE == "a":
                    Predicted = HDriveTime + ParameterDF["alpha"] * HStayTime + (1-ParameterDF["alpha"]) * CStayTime
                elif MODE == "ac":
                    Predicted = HDriveTime + ParameterDF["alpha"] * HStayTime + (1-ParameterDF["alpha"]) * CStayTime + ParameterDF["constant"]
                elif MODE == "ar":
                    Predicted = HDriveTime * Ratio + ParameterDF["alpha"] * HStayTime + (1-ParameterDF["alpha"]) * CStayTime
                elif MODE == "acr":
                    Predicted = HDriveTime * Ratio + ParameterDF["alpha"] * HStayTime + (1-ParameterDF["alpha"]) * CStayTime + ParameterDF["constant"]
                
                Predicted = Predicted.values[0]

                """
                計算指標
                """
                if GroundTruth != 0 and HDriveTime != 0 and (GroundTruth <= avg + std * 3) and (GroundTruth >= avg - std * 3):
                    if PredictedType == "-1":
                        total_1 += 1
                        if abs(Predicted - GroundTruth) <= 10:
                            correct_10_1 += 1
                        if abs(Predicted - GroundTruth) <= 30:
                            correct_30_1 += 1
                        if abs(Predicted - GroundTruth) <= 60:
                            correct_60_1 += 1
                        if abs(Predicted - GroundTruth) <= 120:
                            correct_120_1 += 1

                    elif PredictedType == "-2":
                        total_2 += 1
                        if abs(Predicted - GroundTruth) <= 10:
                            correct_10_2 += 1
                        if abs(Predicted - GroundTruth) <= 30:
                            correct_30_2 += 1
                        if abs(Predicted - GroundTruth) <= 60:
                            correct_60_2 += 1
                        if abs(Predicted - GroundTruth) <= 120:
                            correct_120_2 += 1

                    elif PredictedType == "-other":
                        total_3 += 1
                        if abs(Predicted - GroundTruth) <= 10:
                            correct_10_3 += 1
                        if abs(Predicted - GroundTruth) <= 30:
                            correct_30_3 += 1
                        if abs(Predicted - GroundTruth) <= 60:
                            correct_60_3 += 1
                        if abs(Predicted - GroundTruth) <= 120:
                            correct_120_3 += 1

                ResultWS.append([TEST_DATE, f"{sheet}{PredictedType}", HDriveTime, Ratio, std, HStayTime, CStayTime, GroundTruth, Predicted])

            print(f"{sheet}{PredictedType} 預測完成")
        ResultWB.save(f"inference_result/{ROUTEID}/{DIRECTION}/{MODE}/result.xlsx") 

    StoredResult([total_1, correct_10_1, correct_30_1, correct_60_1, correct_120_1], 
                 [total_2, correct_10_2, correct_30_2, correct_60_2, correct_120_2], 
                 [total_3, correct_10_3, correct_30_3, correct_60_3, correct_120_3])