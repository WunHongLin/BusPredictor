import googlemaps
import argparse
import os
import pandas as pd
import numpy as np
from TimeStastic import RealTimeStastic
from GetInfo import BusSechudleInfo, RouteRealTime
from Tool import FindCarIDIndex, FindClosestTime, CalDiffTime, checkLessThanThreshold
from openpyxl import Workbook, load_workbook

"""
初始化歷史行駛時間表
"""
def InitHistoryDriveTimeTable(ROUTEID, DAY, DIRECTION):
    stastic_driveTime = pd.read_excel(f"StatisticResult/{ROUTEID}/median_drivetime_result_{ROUTEID}_{DAY}_{DIRECTION}.xlsx")
    InitHistoryDriveTimeList = []

    """
    看缺值，如果真的有出現，就用google api 將數值補上
    """
    for index in range(len(stastic_driveTime)):
        certain_sechudle_driveTime = stastic_driveTime.iloc[index].to_list()[1:]
        Lon, Lat = BusSechudleInfo(ROUTEID, DIRECTION, "2025-07-25 00:00:00.000", DEPARTSTOP).GetLonLatInfo()

        for idx in range(len(certain_sechudle_driveTime)): 
            if np.isnan(certain_sechudle_driveTime[idx]):
                departStop = (Lat[idx-1], Lon[idx-1])
                destinationStop = (Lat[idx], Lon[idx])
                """
                下面的金耀之後要改成你的
                """
                gmaps = googlemaps.Client(key="")
                result = gmaps.distance_matrix(departStop, destinationStop, mode="driving", departure_time="now")
                duration = result['rows'][0]['elements'][0]['duration']['value']
                certain_sechudle_driveTime[idx] = duration 

        InitHistoryDriveTimeList.append(certain_sechudle_driveTime)

    return InitHistoryDriveTimeList

"""
判斷進站離站結果是否合法
"""
def isValid(BaseTime, CheckedTimeList, ThresHold_Type):
    """
    這裡是要根據不同的情況去設置不同的閥值
    1. 如果是近程，例如判斷前一站跟當前站 ===> 600s
    2. 如果是遠程，最後一站跟當前站 ===> 3600s
    (數值部分都可以自己在調整看看)
    """
    if ThresHold_Type == 1:
        ThresHold = 600
    else:
        ThresHold = 3600    

    for time in CheckedTimeList:
        if CalDiffTime(BaseTime, time) < ThresHold:
            return True

    return False

"""
取出對應的車次編號
"""
def GetCarID(RealTimeDF, SechudleTime,  ROUTEID, DIRECTION):
    carIDQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == 1) & (RealTimeDF["A2EventType"] == 0)
    carID_DF = RealTimeDF[carIDQuery]
    carIDIndex = FindCarIDIndex(f"{SechudleTime}:00", carID_DF["GPSTime"].apply(str))
    carID = carID_DF.iloc[carIDIndex]["PlateNumb"]
    return carID

"""
取出歷史的統計行駛時間
"""
def GetStasticDriveTime(HistoryDriveTimeTable, CURRENTSTOP, AFTERSTOP, testSechudleIndex):
    certain_sechudle_driveTime = HistoryDriveTimeTable[testSechudleIndex]
    totalDriveTime_of_eachStop = sum(certain_sechudle_driveTime[CURRENTSTOP: AFTERSTOP])

    return totalDriveTime_of_eachStop

"""
取出歷史的統計等待時間
"""
def GetHistoryStayTime(ROUTEID, DAY, DIRECTION, CurrentStopIndex, testSechudleIndex):
    stastic_stayTime = pd.read_excel(f"StatisticResult/{ROUTEID}/median_staytime_result_{ROUTEID}_{DAY}_{DIRECTION}.xlsx")
    history_stasticStayTime = stastic_stayTime.iloc[testSechudleIndex].to_list()[CurrentStopIndex]

    return history_stasticStayTime

"""
取出公車當前站的離站跟前一站的離站的差值
"""
def GetGroundTruth(ROUTEID, DIRECTION, AfterStop, carID, CurrentStopDepartTime, RealTimeDF):
    ArrivalQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == AfterStop) & (RealTimeDF["A2EventType"] == 1) & (RealTimeDF["PlateNumb"] == carID)
    ArrivalResult = RealTimeDF[ArrivalQuery]["GPSTime"].tolist()
    ArrivalResult = [str(time) for time in ArrivalResult]

    if not isValid(CurrentStopDepartTime, ArrivalResult, 2): ArrivalResult = []

    if len(ArrivalResult) != 0:
        ArrivalTime = ArrivalResult[FindClosestTime(CurrentStopDepartTime, ArrivalResult)]
        return CalDiffTime(CurrentStopDepartTime, ArrivalTime)
    else:
        return 0

"""
取出公車當前(GroundTruth)以及歷史行駛時間的比例
"""
def GetRatio(ROUTEID, DIRECTION, CurrentStop, carID, CurrentStopDepartTime, RealTimeDF, testSechudleIndex, HistoryDriveTimeTable):
    """
    首先先取得前一站的離站時間以及當前站的到站時間，求前面一段的行駛時間(GroundTruth)。
    第一站: 則直接回傳比例1
    沒有近離站時間: 同樣回傳比例1
    """
    if CurrentStop == 1: return 1.0, 0, 0

    ArrivalQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == CurrentStop) & (RealTimeDF["A2EventType"] == 1) & (RealTimeDF["PlateNumb"] == carID)
    ArrivalResult = RealTimeDF[ArrivalQuery]["GPSTime"].tolist()
    ArrivalResult = [str(time) for time in ArrivalResult]

    DepartQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == CurrentStop-1) & (RealTimeDF["A2EventType"] == 0) & (RealTimeDF["PlateNumb"] == carID)
    DepartResult = RealTimeDF[DepartQuery]["GPSTime"].tolist()
    DepartResult = [str(time) for time in DepartResult]

    if not isValid(CurrentStopDepartTime, ArrivalResult, 2): ArrivalResult = []
    if not isValid(CurrentStopDepartTime, DepartResult, 2): DepartResult = []

    if len(ArrivalResult) != 0  and len(DepartResult) != 0:
        ArrivalTime = ArrivalResult[FindClosestTime(CurrentStopDepartTime, ArrivalResult)]
        DepartTime = DepartResult[FindClosestTime(CurrentStopDepartTime, DepartResult)]
        GroundTruth = CalDiffTime(DepartTime, ArrivalTime)
    else:
        return 1.0, 0, 0
    
    """
    再來要計算出歷史的行駛時間，可以直接查表
    """
    HistoryDriveTime = GetStasticDriveTime(HistoryDriveTimeTable, CurrentStop-1, CurrentStop, testSechudleIndex)
    
    """
    回傳計算結果
    """
    return GroundTruth / HistoryDriveTime, GroundTruth, HistoryDriveTime

"""
取的對應的範圍
"""
def GetStd(ROUTEID, DAY, DIRECTION, CURRENTSTOP, AFTERSTOP, testSechudleIndex):
    DriveTimeStd = pd.read_excel(f"StatisticResult/{ROUTEID}/std_drivetime_result_{ROUTEID}_{DAY}_{DIRECTION}.xlsx")
    certain_sechudle_driveTime_std = DriveTimeStd.iloc[testSechudleIndex].to_list()[1:]
    totalStd = sum(certain_sechudle_driveTime_std[CURRENTSTOP: AFTERSTOP])

    AvgDriveTime = pd.read_excel(f"StatisticResult/{ROUTEID}/drivetime_result_{ROUTEID}_{DAY}_{DIRECTION}.xlsx")
    certain_sechudle_driveTime_avg = AvgDriveTime.iloc[testSechudleIndex].to_list()[1:]
    totalAvg = sum(certain_sechudle_driveTime_avg[CURRENTSTOP: AFTERSTOP])

    return totalStd, totalAvg, totalAvg + 3 * totalStd, totalAvg - 3 * totalStd

if __name__ == "__main__":
    """
    針對路線、方向、星期幾去整理對應的訓練集
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--routeid", type=int)
    parser.add_argument("--direction", type=int)
    parser.add_argument("--day", type=str)
    opt = parser.parse_args()
    print(opt)

    ROUTEID, DIRECTION, DAY = opt.routeid, opt.direction, opt.day

    """
    這部分程式碼跟StatisticDataPrepare.py 檔案是一樣的，
    1. 先抓取起程站資訊
    2. 根據起程站點將時刻表以及路線資訊取出。
    3. 針對不同的星期，取出整理好的dict
    格式:
    {20251001: df1, 20251008: df2, ...}
    """
    df = pd.read_csv("./Info/v_stg_ibus_dailytimetable.csv")
    query = (df["routeid"] == ROUTEID) & (df["direction"] == DIRECTION)
    DEPARTSTOP = df[query]["departurestopname_zh_tw"].unique()[0]

    if int(DAY) <= 5: busSechudleInfo = BusSechudleInfo(ROUTEID, DIRECTION, "2025-07-25 00:00:00.000", DEPARTSTOP)
    else: busSechudleInfo = BusSechudleInfo(ROUTEID, DIRECTION, "2025-07-26 00:00:00.000", DEPARTSTOP)
    TimeTable = busSechudleInfo.GetTimeTable()
    RouteTable = busSechudleInfo.GetRouteInfo()

    RealTimeDFDict = RouteRealTime(ROUTEID, DIRECTION, DAY, "2025-09-23", "2025-10-29").GetRealTimeDF()

    """
    建立對應資料夾，用於後續的資料新增
    """
    if not os.path.exists(f"training_dataset/{ROUTEID}/{DIRECTION}"): 
        os.makedirs(f"training_dataset/{ROUTEID}/{DIRECTION}")

    """
    由於先前整理好的歷史行駛時間統計表，還是會有缺值(資料真的有夠破)出現，所以先前方法是應用google api 取出站到站的資訊
    在城市一開始，需要先將該表格初始化好，以便後續的資料使用
    """
    HistoryDriveTimeTable = InitHistoryDriveTimeTable(ROUTEID, DAY, DIRECTION)

    """
    初始化一個Excel檔案，用來存放結果
    """
    ResultWB = Workbook()

    for Stop in range(1, len(RouteTable)+1):
        ResultWS = ResultWB.create_sheet(title=f"第{Stop}站")
        ResultWS.append(["Date", "SechudeleIndex", "h_driveTime", "Ratio", "R_gt", "R_ht", "std", "avg", "h_stayTime", "c_stayTime", "GroundTruth"])
        ResultWS.column_dimensions["A"].width = 20
        ResultWS.column_dimensions["B"].width = 20
        ResultWS.column_dimensions["C"].width = 20
        ResultWS.column_dimensions["D"].width = 20
        ResultWS.column_dimensions["E"].width = 20
        ResultWS.column_dimensions["F"].width = 20
        ResultWS.column_dimensions["G"].width = 20
        ResultWS.column_dimensions["H"].width = 20
        ResultWS.column_dimensions["I"].width = 20
        ResultWS.column_dimensions["J"].width = 20
        ResultWS.column_dimensions["K"].width = 20

    ResultWB.save(f"training_dataset/{ROUTEID}/{DIRECTION}/dataset_{DAY}.xlsx")
    
    """
    開始進行資料整理, 迴圈順序為日期 ===> 時刻表 ===> 每一個站點
    """
    for DATE in RealTimeDFDict.keys():
        RealTimeDF = RealTimeDFDict[DATE]
        RealTimeDF["GPSTime"] = pd.to_datetime(RealTimeDF["GPSTime"]).dt.tz_localize(None)
        RealTimeDF["GPSTime"] = RealTimeDF["GPSTime"].dt.time

        for SechudleIndex, SechudleTime in enumerate(TimeTable):
            carID = GetCarID(RealTimeDF, SechudleTime, ROUTEID, DIRECTION)
            PreviousStopDepartTime = f"{SechudleTime}:00"
            """
            將剛剛初始化的exel表格載入
            """
            wb = load_workbook(f"training_dataset/{ROUTEID}/{DIRECTION}/dataset_{DAY}.xlsx")

            for CurrentStop in range(1, len(RouteTable)+1):
                """
                取出當前站的等待時間
                1. 如果是第一站 ===> 等待時間為0
                2. 如果當前站沒有進站或是離站資料 ===> 等待時間為0
                """
                if CurrentStop == 1:
                    CurrentStopStayTime = 0
                else:
                    ArrivalQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == CurrentStop) & (RealTimeDF["A2EventType"] == 1) & (RealTimeDF["PlateNumb"] == carID)
                    ArrivalResult = RealTimeDF[ArrivalQuery]["GPSTime"].tolist()
                    ArrivalResult = [str(time) for time in ArrivalResult]

                    DepartQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == CurrentStop) & (RealTimeDF["A2EventType"] == 0) & (RealTimeDF["PlateNumb"] == carID)
                    DepartResult = RealTimeDF[DepartQuery]["GPSTime"].tolist()
                    DepartResult = [str(time) for time in DepartResult]

                    if not isValid(PreviousStopDepartTime, ArrivalResult, 1): ArrivalResult = []
                    if not isValid(PreviousStopDepartTime, DepartResult, 1): DepartResult = []

                    if len(ArrivalResult) != 0  and len(DepartResult) != 0:
                        ArrivalTime = ArrivalResult[FindClosestTime(PreviousStopDepartTime, ArrivalResult)]
                        DepartTime = DepartResult[FindClosestTime(PreviousStopDepartTime, DepartResult)]
                        CurrentStopStayTime = CalDiffTime(ArrivalTime, DepartTime)
                    else:
                        CurrentStopStayTime = 0

                """
                看當前站有沒有離站時間
                1. 如果沒有則代表沒有辦法跟其他站點計算GroundTruth(進站-離站)，直接跳過迴圈
                """
                DepartQuery = (RealTimeDF["RouteID"] == ROUTEID) & (RealTimeDF["Direction"] == DIRECTION) & (RealTimeDF["StopSequence"] == CurrentStop) & (RealTimeDF["A2EventType"] == 0) & (RealTimeDF["PlateNumb"] == carID)
                DepartResult = RealTimeDF[DepartQuery]["GPSTime"].tolist()
                DepartResult = [str(time) for time in DepartResult]

                if not isValid(PreviousStopDepartTime, DepartResult, 1): DepartResult = []

                if len(DepartResult) != 0:
                    CurrentStopDepartTime = DepartResult[FindClosestTime(PreviousStopDepartTime, DepartResult)]
                    PreviousStopDepartTime = CurrentStopDepartTime
                else:
                    continue
                
                """
                接下來要用一個迴圈去取得後面站點的進站時間(Ground Truth)、歷史行駛時間、API時間、歷史等待時間
                在第一個站點是不需要等待時間的計算的
                """
                ResultWS = wb[f"第{CurrentStop}站"]
                """
                維護兩個陣列，用來儲存累積的當前等待時間以及歷史等待時間
                """
                CurrentStayTimes, StastisticStayTimes = [], []

                for AfterStop in range(CurrentStop+1, len(RouteTable)+1):
                    HistoryDriveTime = GetStasticDriveTime(HistoryDriveTimeTable, CurrentStop, AfterStop, SechudleIndex)
                    Ratio, R_GT, R_HD = GetRatio(ROUTEID, DIRECTION, CurrentStop, carID, CurrentStopDepartTime, RealTimeDF, SechudleIndex, HistoryDriveTimeTable)
                    HistoryStayTime = sum(StastisticStayTimes)
                    CurrentStayTime = sum(CurrentStayTimes)
                    GroundTruth = GetGroundTruth(ROUTEID, DIRECTION, AfterStop, carID, CurrentStopDepartTime, RealTimeDF)
                    """
                    計算完成上述資訊之後，需要將當前預測站的歷史等待時間以及真實等待時間加入到兩個串列之中進行累積
                    """
                    StastisticStayTimes.append(GetHistoryStayTime(ROUTEID, DAY, DIRECTION, AfterStop, SechudleIndex))
                    CurrentStayTimes.append(CurrentStopStayTime)
                    """
                    判斷groundTruth是否為離群直(三倍標準差以外)，如果是的話就跳過本次迴圈
                    """
                    std, avg, UpperBound, LowerBound = GetStd(ROUTEID, DAY, DIRECTION, CurrentStop, AfterStop, SechudleIndex)
                    if GroundTruth > UpperBound or GroundTruth < LowerBound:
                        continue
                    """
                    這裡要根據對應的站點，將資料寫入到對應的Excel活頁中 
                    """
                    if AfterStop - CurrentStop <= 2:
                        ResultWS.append([DATE, f"{SechudleIndex}-{AfterStop-CurrentStop}", HistoryDriveTime, Ratio, R_GT, R_HD, std, avg, HistoryStayTime, CurrentStayTime, GroundTruth])
                    else:
                        ResultWS.append([DATE, f"{SechudleIndex}-other", HistoryDriveTime, Ratio, R_GT, R_HD, std, avg, HistoryStayTime, CurrentStayTime, GroundTruth])

            print(f"日期:{DATE}, 第{SechudleIndex}班次整理完成") 

            wb.save(f"training_dataset/{ROUTEID}/{DIRECTION}/dataset_{DAY}.xlsx")
            
        print(f"{DATE} 儲存完成")