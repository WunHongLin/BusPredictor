import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import numpy as np
import os

"""
計算兩個時間的差距
"""
def CalDiffTime(time1, time2):
    t1 = datetime.strptime(time1, "%H:%M:%S")
    t2 = datetime.strptime(time2, "%H:%M:%S")

    if t1 <= t2: diff = t2 - t1
    else: diff = t1 - t2

    return diff.total_seconds()

"""
找出最近的時間
departTime: str
timeIntervalResult: list[str]
"""
def FindClosestTime(departTime, timeIntervalResult):
    diff = [CalDiffTime(departTime, time) for time in timeIntervalResult]
    return diff.index(min(diff))

"""
找出對應的車次編號
departTime: str
updateTimes: list[str]
"""
def FindCarIDIndex(departTime, updateTimes):
    diff = [CalDiffTime(departTime, updateTime) for updateTime in updateTimes]
    return diff.index(min(diff))

"""
檢查搜尋結果中，是否存在目標時間
BaseTime: str
CheckedTimeList: list[str]
"""
def checkLessThanThreshold(BaseTime, CheckedTimeList):
    result = [CalDiffTime(BaseTime, time)<600 for time in CheckedTimeList]
    """
    如果false數量等同於串列長度，代表所有東西都超出閥值，600===> 10分鐘
    """
    return result.count(False) == len(CheckedTimeList)

"""
將推論結果儲存到excel表中，之後再做簡報會比較方便一些...
"""
def StoredResult(ResultList1, ResultList2, ResultList3):
    total_1, correct_10_1, correct_30_1, correct_60_1, correct_120_1 = ResultList1
    total_2, correct_10_2, correct_30_2, correct_60_2, correct_120_2 = ResultList2
    total_3, correct_10_3, correct_30_3, correct_60_3, correct_120_3 = ResultList3

    if not os.path.exists("inference_acc.xlsx"):
        FileWB = Workbook()
        for index in range(1, 4):
            FileWS = FileWB.create_sheet(title=f"{index}")
            FileWS.append(["total", "acc_10", "acc_30", "acc_60", "acc_120", "acc_others"])
        FileWB.save("inference_acc.xlsx")

    FileWB = load_workbook("inference_acc.xlsx")
    for index, ResultList in enumerate([ResultList1, ResultList2, ResultList3], start=1):
        total, correct_10, correct_30, correct_60, correct_120 = ResultList
        if total == 0: continue
        FileWS = FileWB[f"{index}"]
        Result_10 = f"{round(correct_10/ total, 4)}/ {correct_10}"
        Result_30 = f"{round((correct_30 / total) - (correct_10/ total), 4)}/ {correct_30 - correct_10}"
        Result_60 = f"{round((correct_60 / total) - (correct_30/ total), 4)}/ {correct_60 - correct_30}"
        Result_120 = f"{round((correct_120 / total) - (correct_60/ total), 4)}/ {correct_120 - correct_60}"
        Result_others = f"{round(1 - (correct_120/ total), 4)}/ {total - correct_120}"
        FileWS.append([total, Result_10, Result_30, Result_60, Result_120, Result_others])
        
    FileWB.save("inference_acc.xlsx")
    
    print(f"下一站預測結果：")
    print(f"acc_10_1: {correct_10_1/total_1:.4f}, correct: {correct_10_1}, total: {total_1}")
    print(f"acc_30_1: {(correct_30_1/total_1 - correct_10_1/total_1):.4f}, correct: {correct_30_1 - correct_10_1}, total: {total_1}")
    print(f"acc_60_1: {(correct_60_1/total_1 - correct_30_1/total_1):.4f}, correct: {correct_60_1 - correct_30_1}, total: {total_1}")
    print(f"acc_120_1: {(correct_120_1/total_1 - correct_60_1/total_1):.4f}, correct: {correct_120_1 - correct_60_1}, total: {total_1}")
    print(f"acc_others_1: {1 - (correct_120_1/total_1):.4f}, correct: {total_1 - correct_120_1}, total: {total_1}")

    print(f"下下站預測結果：")
    print(f"acc_10_2: {correct_10_2/total_2:.4f}, correct: {correct_10_2}, total: {total_2}")
    print(f"acc_30_2: {(correct_30_2/total_2 - correct_10_2/total_2):.4f}, correct: {correct_30_2 - correct_10_2}, total: {total_2}")
    print(f"acc_60_2: {(correct_60_2/total_2 - correct_30_2/total_2):.4f}, correct: {correct_60_2 - correct_30_2}, total: {total_2}")
    print(f"acc_120_2: {(correct_120_2/total_2 - correct_60_2/total_2):.4f}, correct: {correct_120_2 - correct_60_2}, total: {total_2}")
    print(f"acc_others_2: {1 - (correct_120_2/total_2):.4f}, correct: {total_2 - correct_120_2}, total: {total_2}")

    print(f"其他站預測結果：")
    print(f"acc_10_3: {correct_10_3/total_3:.4f}, correct: {correct_10_3}, total: {total_3}")
    print(f"acc_30_3: {(correct_30_3/total_3 - correct_10_3/total_3):.4f}, correct: {correct_30_3 - correct_10_3}, total: {total_3}")
    print(f"acc_60_3: {(correct_60_3/total_3 - correct_30_3/total_3):.4f}, correct: {correct_60_3 - correct_30_3}, total: {total_3}")
    print(f"acc_120_3: {(correct_120_3/total_3 - correct_60_3/total_3):.4f}, correct: {correct_120_3 - correct_60_3}, total: {total_3}")
    print(f"acc_others_3: {1 - (correct_120_3/total_3):.4f}, correct: {total_3 - correct_120_3}, total: {total_3}")

"""
因應老師需求，製作的散步圖(主要是看超過120秒在各個區段的數量)
"""
def GenerateGraph(Predicted, GroundTruths):
    diffs = np.abs(np.array(Predicted) - np.array(GroundTruths))
    bins = np.arange(120, max(diffs) + 60, 60)
    plt.figure(figsize=(8, 5))
    plt.hist(diffs, bins=bins, edgecolor='black', alpha=0.7)
    plt.title("Difference Distribution between Predicted and Groundtruths")
    plt.xlabel("Difference Range")
    plt.ylabel("Count")
    plt.xticks(bins)
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    plt.savefig("fig.jpg")